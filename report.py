import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_report(df):
    df['Actual Class'] = df['Actual Class'].astype(str).str.strip().str.lower()
    df['Detected Class'] = df['Detected Class'].astype(str).str.strip().str.lower()

    true_preds = df[df['Actual Class'] == df['Detected Class']]
    true_grouped = true_preds.groupby('Actual Class')['Confidence'].agg(
        True_Pred_Count='count', Max_True='max', Min_True='min', AVG_True='mean'
    ).reset_index()
    true_grouped['AVG_True'] = true_grouped['AVG_True'].round(2)

    false_preds = df[
        (df['Detection Type'].str.lower() == 'false') &
        (df['Detected Class'].str.lower() != 'no detection')
    ]
    false_grouped = false_preds.groupby('Actual Class')['Confidence'].agg(
        False_Pred_Count='count', Max_False='max', Min_False='min', AVG_False='mean'
    ).reset_index()
    false_grouped['AVG_False'] = false_grouped['AVG_False'].round(2)

    merged = pd.merge(true_grouped, false_grouped, on='Actual Class', how='outer').fillna(0)
    merged['True_Pred_Count'] = merged['True_Pred_Count'].astype(int)
    merged['False_Pred_Count'] = merged['False_Pred_Count'].astype(int)
    merged[['Max_True','Min_True','Max_False','Min_False']] = merged[['Max_True','Min_True','Max_False','Min_False']].round(2)

    # Add Total_Count column
    total_count = df.groupby('Actual Class').size().reset_index(name='Total_Count')
    merged = pd.merge(merged, total_count, on='Actual Class', how='left')

    # Capitalize class labels
    merged['Actual Class'] = merged['Actual Class'].str.title()

    # Reorder columns to have 'Total_Count' next to 'Actual Class'
    merged = merged[['Actual Class', 'Total_Count', 'True_Pred_Count', 'Max_True', 'Min_True', 'AVG_True',
                     'False_Pred_Count', 'Max_False', 'Min_False', 'AVG_False']]

    total_row = pd.DataFrame([{
        'Actual Class': 'Total',
        'True_Pred_Count': int(merged['True_Pred_Count'].sum()),
        'Max_True': merged['Max_True'].max(),
        'Min_True': merged['Min_True'].min(),
        'AVG_True': round(true_preds['Confidence'].mean(), 2),
        'False_Pred_Count': int(merged['False_Pred_Count'].sum()),
        'Max_False': merged['Max_False'].max(),
        'Min_False': merged['Min_False'].min(),
        'AVG_False': round(false_preds['Confidence'].mean(), 2),
        'Total_Count': int(df.shape[0])  # Total count of all rows (or you can adjust this logic as needed)
    }])

    final_df = pd.concat([merged, total_row], ignore_index=True)
    return final_df


def style_excel(df):
    output = BytesIO()
    df.to_excel(output, index=False)
    wb = load_workbook(output)
    ws = wb.active

    header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 15

    for row in range(2, ws.max_row + 1):
        is_total_row = ws.cell(row=row, column=1).value == 'Total'
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.border = thin_border
            if is_total_row:
                cell.fill = yellow_fill
                cell.font = bold_font
            elif col in [3, 4, 5, 6]:
                cell.fill = green_fill
            elif col in [7, 8, 9, 10]:
                cell.fill = red_fill

    styled_output = BytesIO()
    wb.save(styled_output)
    styled_output.seek(0)
    return styled_output

# -------------------- STREAMLIT UI -------------------- #
st.set_page_config(page_title="📊 Detection Report Generator", layout="centered")
st.title("📈 YOLOv8 Prediction Report Generator")
st.markdown("Upload your YOLOv8 results file (.csv or .xlsx) to generate a styled performance report.")

uploaded_file = st.file_uploader("Upload Result File", type=["csv", "xlsx"])

if uploaded_file is not None:
    file_type = uploaded_file.name.split('.')[-1]

    if file_type == 'csv':
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    report_df = generate_report(df)
    styled_excel = style_excel(report_df)

    st.success("✅ Report generated successfully!")
    st.dataframe(report_df)

    st.download_button(
        label="📥 Download Styled Report (Excel)",
        data=styled_excel,
        file_name="Styled_True_False_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
